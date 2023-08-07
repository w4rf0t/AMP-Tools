import json
import requests
import openpyxl
import os
import time

W = "\033[0m"
R = "\033[31m"
G = "\033[32m"
O = "\033[33m"
B = "\033[34m"

# def send_file_to_telegram(file_path, file_path1, chat_id, bot_token,target):
#     url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
#     url1 = f"https://api.telegram.org/bot{bot_token}/sendMessage"
#     s = time.time()
#     formatted_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(s))
#     data = {
#         "chat_id": chat_id,
#         "text": f"Kết quả scan {target} vào lúc {formatted_time}"
#     }
#     response0 = requests.post(url1, data=data)
#     with open(file_path, "rb") as file:
#         file1= {"document": file}  # Đọc dữ liệu từ file, không đóng file sau khi đọc
#         response = requests.post(url, data=data, files=file1)

#     with open(file_path1, "rb") as file1:
#         file2= {"document": file1}  # Đọc dữ liệu từ file, không đóng file sau khi đọc
#         response1 = requests.post(url, data=data, files=file2)

def nmap(target, workbook):
    with open(f'Result/{target}/recon/{target}_ip/zoomeye.json', 'r') as f:
        data = json.load(f)
    ip_ports = {}
    if type(data) == list:
        for item in data:
            ip = item['ip']
            if ip not in ip_ports:
                ip_ports[ip] = []
            ip_ports[ip].append(item)
    else:
        ip=data['ip']
        if ip not in ip_ports:
            ip_ports[ip] = []
        ip_ports[ip].append(data)
    worksheet1 = workbook.active
    worksheet1.title = "IP Ports"
    worksheet1.append(('IP', 'Port', 'Title', 'Service',
                      'App', 'Extrainfo', 'Version'))
    for cell in worksheet1[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for ip, ports in ip_ports.items():
        for port in ports:
            row = (ip, port['port'], str(port['title']), port['service'],
                   port['app'], port['extrainfo'], port['version'])
            worksheet1.append(row)




def waf(target, workbook):
    worksheet2 = workbook.create_sheet(title="Firewalls")
    worksheet2.append(('URL', 'Detected', 'Firewall', 'Manufacturer'))
    for cell in worksheet2[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    with open(f'Result/{target}/recon/{target}_waf.json', 'r') as f:
        text = f.read()
        records = text.split('\n][')
        for i, record in enumerate(records, 2):
            fields = json.loads(record.strip('[]\n'))
            url = fields['url']
            detected = fields['detected']
            firewall = fields['firewall']
            manufacturer = fields['manufacturer']
            worksheet2[f'A{i}'] = url
            worksheet2[f'B{i}'] = detected
            worksheet2[f'C{i}'] = firewall
            worksheet2[f'D{i}'] = manufacturer


def hosting(target, workbook):
    try:
        worksheet3 = workbook.create_sheet(title="Hosting")
        worksheet3.append(('IP', 'Domain'))
        for cell in worksheet3[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        with open(f'Result/{target}/recon/ip_to_domain.json', 'r') as f:
            data = json.load(f)
            ip = data['ip']
            domains = data['domain'].split('\n')
            for domain in domains:
                worksheet3.append((ip, domain.strip()))
    except:
        pass


def final(target, workbook):
    with open(f'Result/{target}/recon/final_status_{target}.json', 'r') as f:
        data = json.load(f)
    worksheet4 = workbook.create_sheet(title="Status")
    worksheet4.append(('Website', 'Host', 'Port', 'Scheme', 'Title', 'Tech'))
    for cell in worksheet4[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for item in data:
        website = list(item.keys())[0]
        host = item[website]['host']
        port = item[website]['port']
        scheme = item[website]['scheme']
        title = item[website]['title']
        tech = ', '.join(item[website]['tech'])
        row = (website, host, port, scheme, title, tech)
        worksheet4.append(row)


def subdomains(target, workbook):
    with open(f'Result/{target}/recon/final_subdomain_{target}.txt', 'r') as f:
        text = f.read().splitlines()
    worksheet5 = workbook.create_sheet(title="Final Subdomains")
    worksheet5.append(('Subdomain',))
    for cell in worksheet5[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for domain in text:
        worksheet5.append((domain,))


def live(target, workbook):
    with open(f'Result/{target}/recon/{target}_live.txt', 'r') as f:
        text = f.read().splitlines()
    worksheet6 = workbook.create_sheet(title="Live Domains")
    worksheet6.append(('Live Domain',))
    for cell in worksheet6[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for live_domain in text:
        worksheet6.append((live_domain,))

def dns(target, workbook):
    with open(f'Result/{target}/recon/{target}_dns.json', 'r') as f:
        data = json.load(f)
    # ip_ports = {}
    # if type(data) == list:
    #     for item in data:
    #         ip = item['ip']
    #         if ip not in ip_ports:
    #             ip_ports[ip] = []
    #         ip_ports[ip].append(item)
    # else:
    #     ip=data['ip']
    #     if ip not in ip_ports:
    #         ip_ports[ip] = []
    #     ip_ports[ip].append(data)
    worksheet7 = workbook.active
    worksheet7.title = f"DNS of {target}"
    worksheet7.append(('Arguments', 'Date', 'Type', 'Address', 'Domain', 'MName', 'Recursive', 'Target', 'Exchange'))
    
    for cell in worksheet7[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    for item in data:
        # print(item)
        if 'arguments' in item:
            row = (item.get('arguments', ''), item.get('date', ''), item.get('type', ''),
                   '', '', '', '', '', '')
            worksheet7.append(row)
        elif 'address' in item:
            row = ('', '', '', item.get('address', ''), item.get('domain', ''),
                   item.get('mname', ''), '', '', '')
            worksheet7.append(row)
        elif 'Version' in item:
            row = ('', '', '', item.get('address', ''), item.get('domain', ''),
                   '', item.get('recursive', ''), item.get('target', ''), '')
            worksheet7.append(row)
        elif 'exchange' in item:
            row = ('', '', '', '', item.get('domain', ''),
                   '', '', '', item.get('exchange', ''))
            worksheet7.append(row)
        # print(row)


def exportation_subdomain(target):
    print(B,'[*] Exporting Report...')
    workbook = openpyxl.Workbook()
    nmap(target, workbook)
    waf(target, workbook)
    # final(target, workbook)
    subdomains(target, workbook)
    live(target, workbook)
    dns(target, workbook)
    file_path = f'Result/{target}/recon/Result.xlsx'
    # file_path1 = f'Result/{target}/recon/result.zip'
    workbook.save(file_path)
    # os.remove(f'Result/{target}/recon/{target}_ip/nmap_{target}.json')
    # os.remove(f'Result/{target}/recon/{target}_waf.json')
    # os.remove(f'Result/{target}/recon/final_subdomain_{target}.txt')
    # os.remove(f'Result/{target}/recon/final_status_{target}.json')
    # os.remove(f'Result/{target}/recon/{target}_live.txt')
    # os.system(f"zip -r Result/{target}/recon/result.zip Result/{target}/recon/{target}_url")
    print(G,'[*] Export Completed')
    # print(B,'Sending report....')
    # send_file_to_telegram(file_path, file_path1,'-895049403','6394974317:AAG098D_1b_RgY8EaudD_-_-3i5zG3Zk94c',target)
    # print(G,'Files sent successfully !')
# send_file_to_telegram('Result/etc.vn/recon/Result.xlsx', 'Result/etc.vn/recon/result.zip','-895049403','6394974317:AAG098D_1b_RgY8EaudD_-_-3i5zG3Zk94c')

def exportation_ip(target):
    print(W, 'Exporting Report...')
    workbook = openpyxl.Workbook()
    nmap(target, workbook)
    waf(target, workbook)
    hosting(target, workbook)
    dns(target, workbook)
    file_path = f'Result/{target}/recon/Result.xlsx'
    workbook.save(file_path)
    os.system(f"zip -r Result/{target}/recon/result.zip Result/{target}/recon/{target}_url")
    file_path1 = f'Result/{target}/recon/result.zip'
    # os.remove(f'Result/{target}/recon/{target}_ip/nmap_{target}.json')
    # os.remove(f'Result/{target}/recon/{target}_waf.json')
    # os.remove(f'Result/{target}/recon/ip_to_domain_{target}.json')
    print(G, 'Export Completed')
    print(B,'Sending report....')
    # send_file_to_telegram(file_path, file_path1,'-895049403','6394974317:AAG098D_1b_RgY8EaudD_-_-3i5zG3Zk94c',target)
    print(G,'Files sent successfully !')

